from tkinter import *
import customtkinter
from PIL import Image, ImageTk
import os
import warnings
warnings.filterwarnings('ignore')
customtkinter.set_default_color_theme("blue")
app = customtkinter.CTk()
app.state("zoomed")
app.resizable(False, False)
app.title("Home Page")


def cover_page():
    img1 = ImageTk.PhotoImage(Image.open("COVER_PAGE_IMAGES\\bg2.jpg"))
    l1 = customtkinter.CTkLabel(master=app, image=img1)
    l1.pack()
    frame = customtkinter.CTkFrame(master=l1, width=1020, height=820, corner_radius=15, fg_color="white")
    frame.place(relx=0.5, rely=0.5, anchor=CENTER)
    l2 = customtkinter.CTkLabel(master=frame, text="Registration Management System", font=('arial', 48, "bold"),text_color="#186677")
    l2.place(x=120, y=45)
    login_img = Image.open("COVER_PAGE_IMAGES\\141.jpg")
    resized_login_img = login_img.resize((400, 400))
    img2 = ImageTk.PhotoImage(resized_login_img)
    l2 = customtkinter.CTkLabel(master=frame, image=img2,text="")
    l2.place(x=300, y=150)
    student_login_image = Image.open("COVER_PAGE_IMAGES\\loginn.png")
    # Resize the image
    resized_student_image = student_login_image.resize((90, 90))
    photo1 = ImageTk.PhotoImage(resized_student_image)
    student_button = customtkinter.CTkButton(master=frame, image=photo1, text="Student Portal",fg_color="#d1d2d7",corner_radius=15, font=("arial", 20, "bold"), text_color="black",command=student_portal_app)
    student_button.place(x=370, y=600)
    app.mainloop()



img1 = ImageTk.PhotoImage(Image.open("Images/upload photo.png"))
imageicon3 = ImageTk.PhotoImage(Image.open("Images/search.png"))
imageicon4 = ImageTk.PhotoImage(Image.open("Images/Layer 4.png"))
img = ImageTk.PhotoImage(Image.open("Images/upload photo.png"))
home_icon = ImageTk.PhotoImage(Image.open("COVER_PAGE_IMAGES/home_icon.png"))

def student_portal_app():
    from datetime import date
    import customtkinter
    from tkinter import messagebox
    from tkinter import filedialog
    from PIL import Image, ImageTk
    import os
    from tkinter.ttk import Combobox
    import openpyxl, xlrd
    from openpyxl import Workbook
    import pathlib
     
    background = "#282631"
    framebg = "#EDEDED"
    framefg = "#06283D"
    root = Toplevel(app)
     
     
     # Hide the main page window
    app.withdraw()
    root.title("Student Registration System")
    root.state("zoomed")
    root.resizable(False, False)
    root.geometry("1270x700+210+100")
    root.config(bg=background)
    file = pathlib.Path("Student_data.xlsx")
     
    if file.exists():
        pass
    else:
        file = Workbook()
        sheet = file.active
        sheet['A1'] = "Registration No"
        sheet['B1'] = "Name"
        sheet['C1'] = "Email"
        sheet['D1'] = "Gender"
        sheet['E1'] = "DOB"
        sheet['F1'] = "Date of Registration"
        sheet['G1'] = "Religion"
        sheet['H1'] = "Mobile Number"
        sheet['I1'] = "Father Name"
        sheet['J1'] = "Mother Name"
        sheet['K1'] = "Father's Occupation"
        sheet['L1'] = "Mother's Occupation"
        sheet['M1'] = "Tenth Class Marks"
        sheet['N1'] = "Tenth Passout Year "
        sheet['O1'] = "School"
        sheet['P1'] = "Intermediate Marks"
        sheet['Q1'] = "Inter Passout Year"
        sheet['R1'] = "College"
        file.save('Student_data.xlsx')
    
    def home_page():
      root.withdraw()
      app.deiconify()
    # EXIT
    def Exit():
        root.destroy()
    #######################search#####################3
    def search():
        text = Search.get()
        Clear()
        savebutton.configure(state='disable')
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        if text == "":
           messagebox.showerror("Error","Lack of Information")
        else:
            for row in sheet.rows:
               if row[0].value == int(text):
                   name = row[0]
                   reg_no_position = str(name)[14:-1]
                   reg_number = str(name)[15:-1] 
        try:
           pass
        except:
            messagebox.showerror("Invalid", "Invalid registration nummber!!!")
        x1 = sheet.cell(row=int(reg_number), column=1).value
        x2 = sheet.cell(row=int(reg_number), column=2).value
        x3 = sheet.cell(row=int(reg_number), column=3).value
        x4 = sheet.cell(row=int(reg_number), column=4).value
        x5 = sheet.cell(row=int(reg_number), column=5).value
        x6 = sheet.cell(row=int(reg_number), column=6).value
        x7 = sheet.cell(row=int(reg_number), column=7).value
        x8 = sheet.cell(row=int(reg_number), column=8).value
        x9 = sheet.cell(row=int(reg_number), column=9).value
        x10 = sheet.cell(row=int(reg_number), column=10).value
        x11 = sheet.cell(row=int(reg_number), column=11).value
        x12 = sheet.cell(row=int(reg_number), column=12).value
        x13 = sheet.cell(row=int(reg_number), column=13).value
        x14 = sheet.cell(row=int(reg_number), column=14).value
        x15 = sheet.cell(row=int(reg_number), column=15).value
        x16 = sheet.cell(row=int(reg_number), column=16).value
        x17 = sheet.cell(row=int(reg_number), column=17).value
        x18 = sheet.cell(row=int(reg_number), column=18).value
        Registration.set(x1)
        Name.set(x2)
        Email.set(x3)
        if x4 == 'female':
           R2.select()
        else:
           R1.select()
           DOB.set(x5)
           Date.set(x6)
           Religion.set(x7)
           Mobile_no.set(x8)
           F_Name.set(x9)
           M_Name.set(x10)
           Father_Occupation.set(x11)
           Mother_Occupation.set(x12)
           Tenth_Marks.set(x13)
           Tenth_pass_year.set(x14)
           School_name.set(x15)
           Inter_Marks.set(x16)
           Inter_pass_year.set(x17)
           college_name.set(x18)
           img = (Image.open("Student Images/" + str(x1) + ".jpg"))
           resized_image = img.resize((190, 190))
           photo2 = ImageTk.PhotoImage(resized_image)
           lbl.configure(image=photo2)
           lbl.image = photo2
    ########################### update ###############################
    def Update():
        R1 = Registration.get()
        N1 = Name.get()
        C1 = Email.get()
        selection()
        G1 = gender
        D2 = DOB.get()
        D1 = Date.get()
        Re = Religion.get()
        S1 = Mobile_no.get()
        fathername = F_Name.get()
        mothername = M_Name.get()
        F1 = Father_Occupation.get()
        M1 = Mother_Occupation.get()
        T1 = Tenth_Marks.get()
        P1 = Tenth_pass_year.get()
        Scl = School_name.get()
        I1 = Inter_Marks.get()
        P2 = Inter_pass_year.get()
        Clg = college_name.get()
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        for row in sheet.rows:
            if row[0].value == R1:
                name = row[0]
                reg_no_position = str(name)[14:-1]
                reg_number = str(name)[15:-1]
                # sheet.cell(column=1,row=int(reg_number),value=R1)
                sheet.cell(column=2, row=int(reg_number), value=N1)
                sheet.cell(column=3, row=int(reg_number), value=C1)
                sheet.cell(column=4, row=int(reg_number), value=G1)
                sheet.cell(column=5, row=int(reg_number), value=D2)
                sheet.cell(column=6, row=int(reg_number), value=D1)
                sheet.cell(column=7, row=int(reg_number), value=Re)
                sheet.cell(column=8, row=int(reg_number), value=S1)
                sheet.cell(column=9, row=int(reg_number), value=fathername)
                sheet.cell(column=10, row=int(reg_number), value=mothername)
                sheet.cell(column=11, row=int(reg_number), value=F1)
                sheet.cell(column=12, row=int(reg_number), value=M1)
                sheet.cell(column=13, row=int(reg_number), value=T1)
                sheet.cell(column=14, row=int(reg_number), value=P1)
                sheet.cell(column=15, row=int(reg_number), value=Scl)
                sheet.cell(column=16, row=int(reg_number), value=I1)
                sheet.cell(column=17, row=int(reg_number), value=P2)
                sheet.cell(column=18, row=int(reg_number), value=Clg)
                file.save(r'Student_data.xlsx')
                try:
                    img.save("Student Images/" + str(R1) + ".jpg")
                except:
                    pass
                messagebox.showinfo("Update", "Update Successfully!!")
                
                Clear()
        # upload images
    def showimage():
        global filename
        global img
        filename = filedialog.askopenfilename(initialdir=os.getcwd, title="Select image file",
        filetype=(("Jpg File", "*.jpg"),
                    ("PNG File", "*.png"),
                    ("All Files", "*.txt")))
        img = (Image.open(filename))
        resized_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.configure(image=photo2)
        lbl.image = photo2

    ########################Registration no automatic function#####################
    def registration_no():
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        row = sheet.max_row
        max_row_value = sheet.cell(row=row, column=1).value
        try:
            Registration.set(max_row_value + 1)
        except:
            Registration.set("1")
    ############################# Clear ##################################
    def Clear():
        global img
        global filename
        Name.set("")
        DOB.set("")
        Religion.set("")
        Mobile_no.set("")
        F_Name.set("")
        M_Name.set("")
        Father_Occupation.set("")
        Mother_Occupation.set("")
        Email.set("")
        Tenth_Marks.set("")
        Tenth_pass_year.set("")
        School_name.set("")
        Inter_Marks.set("")
        Inter_pass_year.set("")
        college_name.set("")
        registration_no()
        savebutton.configure(state="normal")
        img1 = PhotoImage(file="Images\\upload photo.png")
        lbl.configure(image=img1)
        lbl.image = img1
        img = ""
        
    ############################# Save ##########################
    def save():
        R1 = Registration.get()
        N1 = Name.get()
        C1 = Email.get()
        try:
            G1 = gender
        except:
            messagebox.showerror("Error", "Select Gender")
        D2 = DOB.get()
        D1 = Date.get()
        Re = Religion.get()
        S1 = Mobile_no.get()
        fathername = F_Name.get()
        mothername = M_Name.get()
        F1 = Father_Occupation.get()
        M1 = Mother_Occupation.get()
        T1 = Tenth_Marks.get()
        P1 = Tenth_pass_year.get()
        Scl = School_name.get()
        I1 = Inter_Marks.get()
        P2 = Inter_pass_year.get()
        Clg = college_name.get()
            
        if N1 == "" or C1 == "Select Class" or D2 == "" or Re == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "" or T1 == "" or P1 == "" or Scl == "" or P2 == "" or I1 == "" or Clg == "":
            messagebox.showerror("Error", "Few Data is Missing")
        else:
            file = openpyxl.load_workbook('Student_data.xlsx')
            sheet = file.active
            sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
            sheet.cell(column=2, row=sheet.max_row, value=N1)
            sheet.cell(column=3, row=sheet.max_row, value=C1)
            sheet.cell(column=4, row=sheet.max_row, value=G1)
            sheet.cell(column=5, row=sheet.max_row, value=D2)
            sheet.cell(column=6, row=sheet.max_row, value=D1)
            sheet.cell(column=7, row=sheet.max_row, value=Re)
            sheet.cell(column=8, row=sheet.max_row, value=S1)
            sheet.cell(column=9, row=sheet.max_row, value=fathername)
            sheet.cell(column=10, row=sheet.max_row, value=mothername)
            sheet.cell(column=11, row=sheet.max_row, value=F1)
            sheet.cell(column=12, row=sheet.max_row, value=M1)
            sheet.cell(column=13, row=sheet.max_row, value=T1)
            sheet.cell(column=14, row=sheet.max_row, value=P1)
            sheet.cell(column=15, row=sheet.max_row, value=Scl)
            sheet.cell(column=16, row=sheet.max_row, value=I1)
            sheet.cell(column=17, row=sheet.max_row, value=P2)
            sheet.cell(column=18, row=sheet.max_row, value=Clg)
            file.save(r'Student_data.xlsx')
            try:
               img.save("Student Images/" + str(R1) + ".jpg")
            except:
               messagebox.showinfo("Info", "Profile Picture is not available !")
            messagebox.showinfo("Info Message", message="Information Saved Successfully")
            Clear()
            registration_no()
     ############################## gender function ######################
    def selection():
        global gender
        value = radio.get()
        if value == 1:
            gender = "Male"
        else:
            gender = "Female"
    # top frames
    Label(root, text="Email: uppadasatwik@gmail.com", width=10, height=3, bg="#2b7a90", anchor="e",font="arial 14 bold").pack(side=TOP, fill=X)
    Label(root, text="STUDENT REGISTRATION PORTAL", width=10, height=2, bg="#539baf", fg="#fff",font="arial 20 bold").pack(side=TOP, fill=X)
    # search box to update
    Search = StringVar()
    customtkinter.CTkEntry(root, textvariable=Search,text_color="black",border_color="#2b7a90",width=240,border_width=3, font=("arial",20),fg_color="white",bg_color="#539baf").place(x=1500, y=88)
    # iageicon3 = PhotoImage(file="Images/search.png")
    Home = customtkinter.CTkButton(root,text="", compound=LEFT, image=home_icon, width=65, bg_color="#539baf",fg_color="#539baf", command=home_page, height=60)
    Home.place(x=10, y=74)
    
    Srch = customtkinter.CTkButton(root, text="Search", compound=LEFT, image=imageicon3, width=65, bg_color="#539baf",fg_color="#539baf", font=("arial", 13 ,"bold"),command=search)
    Srch.place(x=1750, y=84)
    # iageicon4 = PhotoImage(file="Images/Layer 4.png")
    update_button = customtkinter.CTkButton(root,text="", image=imageicon4, fg_color="#539baf",command=Update,bg_color="#539baf",width=65)
    update_button.place(x=110, y=80)

    # Rgistation and date
    customtkinter.CTkLabel(root, text="Registration No:", font=("arial" ,12,"bold"),fg_color=background).place(x=60, y=150)
    customtkinter.CTkLabel(root, text="Date:", font=("arial ",12,"bold"), fg_color=background,).place(x=1650,y=150)
    Registration = IntVar()
    Date = StringVar()
    # Rg_no
    reg_entry = customtkinter.CTkEntry(root, textvariable=Registration, width=150,font=("arial",12),fg_color="white",text_color="black")
    reg_entry.place(x=160, y=152)
    registration_no()
    # dte
    today = date.today()
    d1 = today.strftime("%d/%m/%Y")
    date_entry = customtkinter.CTkEntry(root, textvariable=Date, width=150, font=("arial",12),fg_color="white",text_color="black")
    date_entry.place(x=1700, y=152)
    Date.set(d1)
    
    # Sudent Details
    obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=1500, bg=framebg, height=250,relief=GROOVE)
    obj.place(x=60, y=200)
    Label(obj, text="Full Name:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=50)
    Label(obj, text="Date of Birth:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=100)
    Label(obj, text="Gender:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=150)
    Label(obj, text="Email:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=50)
    Label(obj, text="Religion:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=100)
    Label(obj, text="Mobile.No:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=150)
    # etries
    Name = StringVar()
    name_entry = customtkinter.CTkEntry(obj, textvariable=Name, width=200, font=("arial ", 16),text_color="black",fg_color="white")
    name_entry.place(x=240, y=50)
    DOB = StringVar()
    dob_entry = customtkinter.CTkEntry(obj, textvariable=DOB, width=200, font=("arial ", 16), text_color="black",fg_color="white")
    dob_entry.place(x=240, y=100)
        
    radio = IntVar()
    R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection,font="arial 16")
    R1.place(x=240, y=150)
    R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection,font="arial 16")
    R2.place(x=330, y=150)
    
    Religion = StringVar()
    rel_entry = customtkinter.CTkEntry(obj, textvariable=Religion, width=200, font=("arial ", 16), text_color="black",fg_color="white")
    rel_entry.place(x=1230, y=100)
    Mobile_no = StringVar()
    skill_entry = customtkinter.CTkEntry(obj, textvariable=Mobile_no, width=200, font=("arial ", 16),text_color="black", fg_color="white")
    skill_entry.place(x=1230, y=150)
    Email = StringVar()
    Email_entry = customtkinter.CTkEntry(obj, textvariable=Email, width=200, font=("arial ", 16),text_color="black",fg_color="white")
    Email_entry.place(x=1230, y=50)
    # Cass = customtkinter.CTkComboBox(obj, values=['1', '2', '3', '4', '5', '7', '8', '9', '10', '11', '12'], font=("Roboto",16), width=200,text_color="black",fg_color="white",bg_color="black"
    # ,tate='r')
    # Cass.place(x=1230, y=50)
    # Cass.set("Select Class")
    # Prent Details
    obj2 = LabelFrame(root, text="Parent's Details", font="48", bd=2, width=1500, bg=framebg, height=220,relief=GROOVE)
    obj2.place(x=60, y=470)
    
    Label(obj2, text="Father's Name:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=50)
    Label(obj2, text="Occupation:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=100)
    F_Name = StringVar()
    f_entry = customtkinter.CTkEntry(obj2, textvariable=F_Name, width=200, font=("arial ", 16), text_color="black",fg_color="white")
    f_entry.place(x=240, y=50)
    Father_Occupation = StringVar()
    FO_Entry = customtkinter.CTkEntry(obj2, textvariable=Father_Occupation, width=200, font=("arial ", 16),text_color="black", fg_color="white")
    FO_Entry.place(x=240, y=100)
    Label(obj2, text="Mother's Name:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=50)
    Label(obj2, text="Occupation:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=100)
    

    M_Name = StringVar()
    m_entry = customtkinter.CTkEntry(obj2, textvariable=M_Name, width=200, font=("arial ", 16), text_color="black",fg_color="white")
    m_entry.place(x=1230, y=50)
    Mother_Occupation = StringVar()
    MO_Entry = customtkinter.CTkEntry(obj2, textvariable=Mother_Occupation, width=200, font=("arial ", 16),text_color="black", fg_color="white")
    MO_Entry.place(x=1230, y=100)
    # iage
    f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
    f.place(x=1650, y=200)
    obj3 = LabelFrame(root, text="Acadamic Details", font="48", bd=2, width=1500, bg=framebg, height=220,relief=GROOVE)
    obj3.place(x=60, y=710)
    Label(obj3, text="Tenth Marks:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=50)
    Label(obj3, text="Tenth Pass Out Year:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=100)
    Label(obj3, text="School Name:", font="arial 16", bg=framebg, fg=framefg).place(x=30, y=150)
    Label(obj3, text="Inter Marks:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=50)
    Label(obj3, text="Inter Pass Out Year:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=100)
    Label(obj3, text="College Name:", font="arial 16", bg=framebg, fg=framefg).place(x=1030, y=150)
    # etries
    Tenth_Marks = StringVar()
    Tenth_Marks_entry = customtkinter.CTkEntry(obj3, width=200, textvariable=Tenth_Marks, fg_color="white",text_color="black", font=("arial", 16))
    Tenth_Marks_entry.place(x=240, y=50)
    Tenth_pass_year = StringVar()
    Tenth_pass_year_entry = customtkinter.CTkEntry(obj3, textvariable=Tenth_pass_year, width=200, font=("arial ", 16),text_color="black", fg_color="white")
    Tenth_pass_year_entry.place(x=240, y=100)
    School_name = StringVar()
    School_name_entry = customtkinter.CTkEntry(obj3, textvariable=School_name, width=200, font=("arial ", 16),text_color="black", fg_color="white")
    School_name_entry.place(x=240, y=150)
    Inter_Marks = StringVar()
    Inter_Marks_entry = customtkinter.CTkEntry(obj3, textvariable=Inter_Marks, width=200, font=("arial ", 16),
                                               text_color="black", fg_color="white")
    Inter_Marks_entry.place(x=1230, y=50)
    Inter_pass_year = StringVar()
    Inter_pass_year_entry = customtkinter.CTkEntry(obj3, textvariable=Inter_pass_year, width=200, font=("arial ", 16), text_color="black", fg_color="white")
    Inter_pass_year_entry.place(x=1230, y=100)

    college_name = StringVar()
    college_name_entry = customtkinter.CTkEntry(obj3, textvariable=college_name, width=200, font=("arial ", 16),
    text_color="black", fg_color="white")
    college_name_entry.place(x=1230, y=150)
     # mg = PhotoImage(file="Images/upload photo.png")
    lbl = Label(f, bg="black", image=img)
    lbl.place(x=0, y=0)
     # utton
     # utton(root,text="Take Picture",width=10,height=2,font=("arial 12 bold"),bg="lightyellow").place(x=977,y=370)
    customtkinter.CTkButton(root, text="Upload", width=180, text_color="black", height=50, font=("arial", 16,"bold"),fg_color="pink", bg_color="#282631", command=showimage).place(x=1650, y=450)
    savebutton = customtkinter.CTkButton(root, text="Save", width=180, text_color="black", height=50,font=("arial", 16, "bold"), fg_color="light green", bg_color="#282631",command=save)
    savebutton.place(x=1650, y=530)
    customtkinter.CTkButton(root, text="Reset", width=180, text_color="black", height=50, font=("arial", 16,"bold"),fg_color="orange", bg_color="#282631", command=Clear).place(x=1650, y=610)
    customtkinter.CTkButton(root, text="Exit", width=180, text_color="black", height=50, font=("arial", 16,"bold"),fg_color="crimson", bg_color="#282631", cursor="pirate", command=Exit).place(x=1650,y=690)
    root.mainloop()
    
cover_page()